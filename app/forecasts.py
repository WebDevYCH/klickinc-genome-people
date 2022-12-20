import datetime, json, os, re
from dateutil.relativedelta import relativedelta
import pandas as pd

from flask import render_template, request
from flask_login import login_required
from flask_admin import expose

from core import *
from model import *

from google.cloud import bigquery
import openpyxl

###################################################################
## MODEL

PortfolioForecast = Base.classes.portfolio_forecast

PortfolioLRForecast = Base.classes.portfolio_laborrole_forecast

PortfolioLRForecastSheet = Base.classes.portfolio_laborrole_forecast_sheet

LaborRoleHoursDayRatio = Base.classes.labor_role_hours_day_ratio

admin.add_view(AdminModelView(PortfolioLRForecastSheet, db.session, category='Forecasts'))
admin.add_view(AdminModelView(LaborRoleHoursDayRatio, db.session, category='Forecasts'))

###################################################################
## UTILITIES

def queryClientCst(clients, csts):
    if csts != None and csts != "":
        queryfilter = Portfolio.currcst.in_(csts.split(','))
    elif clients != None and clients != "":
        queryfilter = Portfolio.clientid.in_(clients.split(','))
    else:
        queryfilter = True
    return queryfilter

# get the portfolio forecasts (in dollars), returns a dictionary
def get_pfs(year, clients, csts, doforecasts=True, doactuals=True, dotargets=True):
    queryfilter = queryClientCst(clients, csts)

    pfs = db.session.query(PortfolioForecast).\
        join(Portfolio).\
        filter(PortfolioForecast.yearmonth >= datetime.date(year,1,1),\
            PortfolioForecast.yearmonth < datetime.date(year+1,1,1),\
            queryfilter,\
        ).\
        all()

    bypfid= {}
    for pf in pfs:
        # clients (as parent nodes)
        key = f"{pf.portfolio.clientid}"
        pfout = bypfid.get(key) or {}
        pfout['id'] = key
        pfout['name'] = pf.portfolio.clientname
        bypfid[key] = pfout

        # portfolios with forecasts
        if doforecasts:
            key = f"{pf.portfolioid}"
            pfout = bypfid.get(key) or {}
            pfout['id'] = key
            pfout['parent'] = f"{pf.portfolio.clientid}"
            pfout['name'] = pf.portfolio.name
            if pf.forecast != None:
                pfout[f"m{pf.yearmonth.month}"] = re.sub('\\...$','',pf.forecast)
                bypfid[key] = pfout

        # targets (as child nodes)
        if dotargets:
            key = f"t{pf.portfolioid}"
            pfout = bypfid.get(key) or {}
            pfout['id'] = key
            pfout['parent'] = f"{pf.portfolioid}"
            pfout['name'] = "Target"
            if pf.target != None:
                pfout[f"m{pf.yearmonth.month}"] = re.sub('\\...$','',pf.target)
                bypfid[key] = pfout

        # actuals (as child nodes)
        if doactuals:
            key = f"a{pf.portfolioid}"
            pfout = bypfid.get(key) or {}
            pfout['id'] = key
            pfout['parent'] = f"{pf.portfolioid}"
            pfout['name'] = "Actuals"
            if pf.actuals != None:
                pfout[f"m{pf.yearmonth.month}"] = re.sub('\\...$','',pf.actuals)
                bypfid[key] = pfout
    return bypfid

# get the portfolio labor role forecasts (in hours), returns a dictionary
def get_plrfs(year, clients, csts, showsources=True):
    queryfilter = queryClientCst(clients, csts)

    pflrs = db.session.query(PortfolioLRForecast).\
        join(Portfolio).\
        join(LaborRole).\
        filter(PortfolioLRForecast.yearmonth >= datetime.date(year,1,1),
            PortfolioLRForecast.yearmonth < datetime.date(year+1,1,1),
            queryfilter
        ).all()

    bypfid = {}
    for pflr in pflrs:
        # add the labor category under the portfolio
        catkey = f"LRCAT-{pflr.portfolioid}-"+re.sub('[^a-zA-Z0-9]','_',pflr.labor_role.categoryname)
        pfout = bypfid.get(catkey) or {}
        pfout['id'] = catkey
        pfout['parent'] = f"{pflr.portfolioid}"
        pfout['name'] = pflr.labor_role.categoryname
        bypfid[catkey] = pfout

        # add the labor role under the category, with the "main" forecast or "linear", with main taking precedence
        forecastkey = f"LR-(LRCAT{pflr.labor_role.categoryname})-{pflr.portfolioid}-{pflr.laborroleid}-MAIN"
        if pflr.forecastedhours != None :
            pfout = bypfid.get(forecastkey) or {}
            pfout['id'] = forecastkey
            pfout['parent'] = catkey
            pfout['name'] = pflr.labor_role.name
            pfout['lrname'] = pflr.labor_role.name
            pfout['source'] = 'MAIN'
            monthkey = f"m{pflr.yearmonth.month}"
            if pflr.source == 'MAIN' or (monthkey not in pfout and pflr.source == 'linear'):
                # TODO: have some hierarchy of the sources
                pfout[f"m{pflr.yearmonth.month}"] = pflr.forecastedhours
                pfout[f"m{pflr.yearmonth.month}-source"] = pflr.source
            bypfid[forecastkey] = pfout

        # if this isn't the main source, add a line underneath that gives the source's forecast
        if showsources and pflr.source != 'MAIN':
            fsourcekey = f"LR-(LRCAT{pflr.labor_role.categoryname})-{pflr.portfolioid}-{pflr.laborroleid}-{pflr.source}"
            pfout = bypfid.get(fsourcekey) or {}
            pfout['id'] = fsourcekey
            pfout['parent'] = forecastkey
            pfout['name'] = f"'{pflr.source}' Forecast"
            pfout['lrname'] = pflr.labor_role.name
            pfout['source'] = pflr.source
            if pflr.forecastedhours != None:
                pfout[f"m{pflr.yearmonth.month}"] = pflr.forecastedhours
            bypfid[fsourcekey] = pfout

    return bypfid

# get the department labor role forecasts (in hours), returns a dictionary
def get_dlrfs(year, lrcat, clients = None, csts = None, showportfolios=True, showsources=True):
    queryfilter = queryClientCst(clients, csts)

    # create a pandas dataframe of the portfolio labor role forecasts
    header = ['id','parent','name','detail',
    'm1','m2','m3','m4','m5','m6','m7','m8','m9','m10','m11','m12',
    'm1src','m2src','m3src','m4src','m5src','m6src','m7src','m8src','m9src','m10src','m11src','m12src']
    df = pd.DataFrame(columns=header)

    # get the portfolio labor role forecasts, querying by lrcat and optionally by client or cst
    if csts != None and csts != "":
        queryfilter = Portfolio.currcst.in_(csts.split(','))
    elif clients != None and clients != "":
        queryfilter = Portfolio.clientid.in_(clients.split(','))
    else:
        queryfilter = True

    # query the labor role forecasts
    pflrs = db.session.query(PortfolioLRForecast).join(Portfolio).join(LaborRole).filter(
            PortfolioLRForecast.yearmonth >= datetime.date(year,1,1),
            PortfolioLRForecast.yearmonth < datetime.date(year+1,1,1),
            LaborRole.categoryname == lrcat,
            queryfilter
        ).all()

    # add the portfolio labor role forecasts to the dataframe
    for pflr in pflrs:
        # we want the hierarchy to be: labor role -> portfolio -> source

        # add a row for the labor role as a root node, if it isn't already there
        lrmainkey = f"{pflr.laborroleid}"
        if lrmainkey not in df.id.values:
            df = pd.concat([df, 
                pd.DataFrame({ 'id': lrmainkey,
                'parent': None,
                'name': pflr.labor_role.name }, index=[0])])
        
        # add a row for the sum of the forecasted hours for the labor role by portfolio
        lrportfoliokey = f"{pflr.portfolioid}-{pflr.laborroleid}"
        if lrportfoliokey not in df.id.values:
            df = pd.concat([df, 
                pd.DataFrame({ 'id': lrportfoliokey,
                'parent': lrmainkey,
                'name': f"{pflr.portfolio.clientname} - {pflr.portfolio.name}",
                'detail': 'portfolio' }, index=[0])])

        # add a row for the sum of the forecasted hours for the labor role by portfolio and source
        lrsourcekey = f"{pflr.portfolioid}-{pflr.laborroleid}-{pflr.source}"
        if lrsourcekey not in df.id.values:
            df = pd.concat([df,
                pd.DataFrame({ 'id': lrsourcekey,
                'parent': lrportfoliokey,
                'name': f"'{pflr.source}' Forecast",
                'detail': 'source' }, index=[0])])

        # fill in the month columns in the source row
        if pflr.forecastedhours != None and pflr.forecastedhours != 0:
            df.loc[df['id'] == lrsourcekey, f"m{pflr.yearmonth.month}"] = pflr.forecastedhours
            df.loc[df['id'] == lrsourcekey, f"m{pflr.yearmonth.month}src"] = pflr.source

    # second pass: fill in portfolio and labor role totals
    # TODO: finish this

    # remove the portfolio and source rows if we don't want them
    if not showportfolios:
        df = df[df['detail'] != 'portfolio']
    if not showsources:
        df = df[df['detail'] != 'source']

    # switch dataframe nulls to blank
    df = df.fillna('')

    return df

# get the current rate for each client per labor role (in dollars)
# returns a dictionary of dictionaries, client -> laborroleid -> rate
def get_clrrates():
    # query from Genome Reports
    #https://genome.klick.com/querytemplate/report-tool.html#/templateID/2093?nocache=true
    json = retrieveGenomeReport(2093)

    rates = {}
    for row in json['Entries']:
        clientrate = rates.get(row['Client']) or {}
        clientrate[row['LaborRoleID']] = row['PerHour']
        rates[row['Client']] = clientrate

    return rates

###################################################################
## LABOR FORECASTS FRONTEND PAGES

# GET /forecasts/portfolio-forecasts
@app.route('/forecasts/portfolio-forecasts')
@login_required
def portfolio_forecasts():
    thisyear = datetime.date.today().year
    # if it's after October, show the next year's forecasts
    if datetime.date.today().month > 10:
        thisyear += 1

    startyear = thisyear-2
    endyear = thisyear+1
    return render_template('forecasts/portfolio-forecasts.html', 
        title='Portfolio Forecasts', 
        startyear=startyear, endyear=endyear, thisyear=thisyear)

# GET /forecasts/portfolio-lr-forecasts
@app.route('/forecasts/portfolio-lr-forecasts')
@login_required
def portfolio_lr_forecasts():
    thisyear = datetime.date.today().year
    # if it's after October, show the next year's forecasts
    if datetime.date.today().month > 10:
        thisyear += 1

    startyear = thisyear-2
    endyear = thisyear+1
    return render_template('forecasts/portfolio-lr-forecasts.html', 
        title='Labor Role Forecasts by Portfolio', 
        startyear=startyear, endyear=endyear, thisyear=thisyear)

# GET /forecasts/dept-lr-forecasts
@app.route('/forecasts/dept-lr-forecasts')
@login_required
def dept_lr_forecasts():
    thisyear = datetime.date.today().year
    # if it's after October, show the next year's forecasts
    if datetime.date.today().month > 10:
        thisyear += 1

    startyear = thisyear-2
    endyear = thisyear+1
    return render_template('forecasts/dept-lr-forecasts.html', 
        title='Labor Role Forecasts', 
        startyear=startyear, endyear=endyear, thisyear=thisyear)

# GET /forecasts/portfolio-forecasts-data
@app.route('/forecasts/portfolio-forecasts-data')
@login_required
def portfolio_forecasts_data():
    year = int(request.args.get('year'))
    clients = request.args.get('clients')
    csts = request.args.get('csts')

    bypfid = get_pfs(year, clients, csts)
    app.logger.info(f"portfolio_forecasts_data size: {len(bypfid)}")

    retval = list(bypfid.values())
    retval.sort(key=lambda x: x['name'])
    return retval

# GET /forecasts/portfolio-lr-forecasts-data
@app.route('/forecasts/portfolio-lr-forecasts-data')
@login_required
def portfolio_forecasts_lr_data():
    year = int(request.args.get('year'))
    clients = request.args.get('clients')
    csts = request.args.get('csts')

    bypfid = get_pfs(year, clients, csts, doactuals=False, dotargets=False) | get_plrfs(year, clients, csts, showsources=True)
    app.logger.info(f"portfolio_lr_forecasts_data size: {len(bypfid)}")
    if len(bypfid) > 2000:
        bypfid = get_pfs(year, clients, csts, doactuals=False, dotargets=False) | get_plrfs(year, clients, csts, showsources=False)
        app.logger.info(f"portfolio_lr_forecasts_data minimized size: {len(bypfid)}")

    retval = list(bypfid.values())
    retval.sort(key=lambda x: x['name'])

    return retval

# GET /forecasts/dept-lr-forecasts-data
@app.route('/forecasts/dept-lr-forecasts-data')
@login_required
def dept_lr_forecasts_data():
    year = int(request.args.get('year'))
    clients = request.args.get('clients')
    csts = request.args.get('csts')
    lrcat = request.args.get('lrcat')
    showportfolios = request.args.get('showportfolios') == 'true'
    showsources = request.args.get('showsources') == 'true'

    # cache the results for 5 minutes
    df = Cache.get(f"dept_lr_forecasts_data_{year}_{clients}_{csts}_{lrcat}_{showportfolios}_{showsources}")
    if df is None:
        df = get_dlrfs(year, lrcat, clients, csts, showportfolios, showsources)
        app.logger.info(f"dept_lr_forecasts_data size: {len(df)}")
        df.sort_values(by=['name'], inplace=True)
        app.logger.info(f"dept_lr_forecasts_data sorted size: {len(df)}")
        Cache.set(f"dept_lr_forecasts_data_{year}_{clients}_{csts}_{lrcat}_{showportfolios}_{showsources}", df, timeout=300)

    retval = df.to_dict('records')
    # remove any null values for the parent column
    for r in retval:
        if r['parent'] == None or r['parent'] == "":
            r.pop('parent')

    return retval


# GET /forecasts/client-list
@app.route('/forecasts/client-list')
@login_required
def pf_client_list():
    year = int(request.args.get('year'))
    clients = db.session.query(Portfolio).\
        distinct(Portfolio.clientid,Portfolio.clientname).\
        join(PortfolioForecast).\
        filter(PortfolioForecast.yearmonth >= datetime.date(year,1,1)).\
        filter(PortfolioForecast.yearmonth < datetime.date(year+1,1,1)).\
        order_by(Portfolio.clientname).\
        all()

    return [{"id":c.clientid, "value":c.clientname } for c in clients]

# GET /forecasts/cst-list
@app.route('/forecasts/cst-list')
@login_required
def pf_cst_list():
    year = int(request.args.get('year'))
    csts = db.session.query(Portfolio).\
        distinct(Portfolio.currcst).\
        join(PortfolioForecast).\
        filter(PortfolioForecast.yearmonth >= datetime.date(year,1,1)).\
        filter(PortfolioForecast.yearmonth < datetime.date(year+1,1,1)).\
        order_by(Portfolio.currcst).\
        all()

    return [{"id":c.currcst, "value":c.currcst } for c in csts]

# GET /forecasts/lrcat-list
@app.route('/forecasts/lrcat-list')
@login_required
def pf_lrcat_list():
    year = int(request.args.get('year'))
    lrcats = db.session.query(LaborRole).\
        distinct(LaborRole.categoryname).\
        join(PortfolioLRForecast).\
        filter(PortfolioLRForecast.yearmonth >= datetime.date(year,1,1)).\
        filter(PortfolioLRForecast.yearmonth < datetime.date(year+1,1,1)).\
        order_by(LaborRole.categoryname).\
        all()

    return [{"id":c.categoryname, "value":c.categoryname } for c in lrcats]

# POST /forecasts/portfolio-forecast-save
@app.post('/forecasts/portfolio-lr-forecast-save')
@login_required
def portfolio_lr_forecast_save():
    data = request.get_json()
    app.logger.info(f"portfolio_lr_forecast_save: {data}")
    # id is based on forecastkey = f"{pflr.portfolioid}-{pflr.laborroleid}-MAIN"
    portfolioid = int(data['id'].split('-')[1])
    laborroleid = data['id'].split('-')[2]
    source = data['id'].split('-')[3]
    if source != 'MAIN':
        return "ERROR: Only MAIN source is allowed"
    year = int(data['year'])
    month = int(data['month'])
    yearmonth = datetime.date(year, month, 1)

    pflr = db.session.query(PortfolioLRForecast).filter_by(portfolioid=portfolioid, laborroleid=laborroleid, yearmonth=yearmonth, source=source).first()
    if data['value'] == '':
        # blank value means they're removing the override (0 to override with 0 hours)
        if pflr:
            db.session.delete(pflr)
        db.session.commit()
        return "OK"
    else:
        # non-blank value means they're setting or changing the override
        if not pflr:
            pflr = PortfolioLRForecast(portfolioid=portfolioid, laborroleid=laborroleid, yearmonth=yearmonth, source='MAIN')
            db.session.add(pflr)
        pflr.forecastedhours = float(data['value'])
        pflr.updateddate = datetime.datetime.now()
        pflr.userid = current_user.userid
        db.session.commit()
    return "OK"

###################################################################
## ADMIN PAGES AND FORECASTING ALGORITHMS

class ForecastAdminView(AdminBaseView):
    @expose('/')
    def index(self):
        pages = {
            'linear': 'Linear Extrapolation Model',
            'cilinear': 'CI + Linear Extrapolation Model',
            'gsheets': 'Google Sheets Forecasts Import',
            'lr_hours_day_ratio': 'Labor Role Hours/Day Ratio Import'
        }
        return self.render('admin/job_index.html', title="Resource Forecast Processing", pages=pages)

    @expose('/linear')
    def linear(self):
        return self.render('admin/job_log.html', loglines=model_linear())

    @expose('/cilinear')
    def cilinear(self):
        return self.render('admin/job_log.html', loglines=model_cilinear())

    @expose('/gsheets')
    def gsheets(self):
        return self.render('admin/job_log.html', loglines=forecast_gsheets())

    @expose('/lr_hours_day_ratio')
    def lr_hours_day_ratio(self):
        return self.render('admin/job_log.html', loglines=replicate_labor_role_hours_day_ratio())

@app.cli.command('model_linear')
def model_linear_cmd():
    model_linear()

def model_linear():
    loglines = AdminLog()
    with app.app_context():
        Base.prepare(autoload_with=db.engine, reflect=True)
    loglines.append("Starting Same-Portfolio Linear Extrapolator")
    loglines.append("")

    lookback = 4
    lookahead = 4
    sourcename = 'linear'
    startdate = datetime.date.today().replace(day=1)

    # for each portfolio with forecasts in the next lookahead months
    loglines.append("grabbing portfolio list")
    for p in db.session.query(PortfolioForecast).where(
                PortfolioForecast.yearmonth >= startdate,
                PortfolioForecast.forecast != None,
                PortfolioForecast.forecast != "0.00"
            ).distinct(PortfolioForecast.portfolioid).all():
        loglines.append(f"portfolio {p.portfolioid}")
        # pick up the forecasts from the Genome report (hardcoded numbers come from the report config)
        json = retrieveGenomeReport(2145, [2434,2435,2436], [p.portfolioid,lookback,lookahead])
        loglines.append(f"{json}")

        if 'Entries' in json:
            # for each Genome forecast
            for pfin in json['Entries']:
                # {'AccountPortfolioID': 174, 'YearMonth': '/Date(1669870800000-0500)/', 'Forecast': 54000.0, 'LaborCategory': 'Analytics', 'LaborRoleID': 'ANLTCADR', 'LaborRoleName': 'Analytics, Associate Director', 'PredictedHour': 0.14, 'PredictedAmount': 17.2267}
                pfin['YearMonth'] = parseGenomeDate(pfin['YearMonth'])
                key = f"{pfin['AccountPortfolioID']} // {pfin['YearMonth']} // {pfin['LaborRoleID']} // {sourcename}"

                if pfin['PredictedHour'] != None:
                    pflr = db.session.query(PortfolioLRForecast).filter(
                        PortfolioLRForecast.portfolioid == pfin['AccountPortfolioID'],
                        PortfolioLRForecast.yearmonth == pfin['YearMonth'],
                        PortfolioLRForecast.laborroleid == pfin['LaborRoleID'],
                        PortfolioLRForecast.source == sourcename).first()

                    if pflr != None and pflr.forecastedhours == pfin['PredictedHour']:
                        loglines.append(f"  SKIPPED {key}")
                    else:
                        if pflr != None:
                            loglines.append(f"  UPDATING {key} because not: {pflr.forecastedhours} == {pfin['PredictedHour']} and {pflr.forecasteddollars} == {pfin['PredictedAmount']}")
                        else:
                            loglines.append(f"  NEW {key}")
                            pflr = PortfolioLRForecast()
                            db.session.add(pflr)

                        # set the fields
                        pflr.portfolioid = pfin['AccountPortfolioID']
                        pflr.yearmonth = pfin['YearMonth']
                        pflr.laborroleid = pfin['LaborRoleID']
                        pflr.source = sourcename
                        pflr.userid = None
                        pflr.updateddate = datetime.date.today()
                        pflr.forecasteddollars = pfin['PredictedAmount']
                        pflr.forecastedhours = pfin['PredictedHour']
        else:
            loglines.append("ERROR: CRASH RETRIEVING PORTFOLIO'S FORECASTS")

        db.session.commit()

    return loglines

@app.cli.command('model_cilinear')
def model_cilinear_cmd():
    model_cilinear()

def model_cilinear():
    loglines = AdminLog()
    with app.app_context():
        Base.prepare(autoload_with=db.engine, reflect=True)
    loglines.append("Starting CI-Selected-Portfolio Linear Extrapolator")
    loglines.append("")

    lookahead = 4
    sourcename = 'cilinear'
    startdate = datetime.date.today().replace(day=1)
    enddate = startdate + relativedelta(months=lookahead)

    # gather the labor rates
    loglines.append("gathering labor rates")
    rates = get_clrrates()

    loglines.append("querying Genome BQ model")
    # query the CI-selected portfolios view in BigQuery
    # query returns AccountPortfolioID, LaborRoleID, Pct, CST, Client
    client = bigquery.Client()
    query_job = client.query("SELECT * FROM `genome-datalake-prod.GenomeReports.vDemandPlanningByPortfolio`")
    results = query_job.result()
    # convert the results to a dictionary by portfolio ID with an array of labor role IDs and percentages 
    ciportfolios = {}
    for row in results:
        ciportfolios[row['AccountPortfolioID']] = ciportfolios.get(row['AccountPortfolioID']) or []
        ciportfolios[row['AccountPortfolioID']].append(row)

    # grab each future forecast
    loglines.append("grabbing portfolio forecast list")
    # for each future forecast
    for pf in db.session.query(PortfolioForecast).where(
                PortfolioForecast.yearmonth >= startdate,
                PortfolioForecast.yearmonth < enddate,
            ).join(Portfolio).all():
        loglines.append(f"portfolio {pf.portfolioid} forecast at {pf.yearmonth}, {pf.forecast}")

        # delete existing portfolio labor role forecasts
        db.session.query(PortfolioLRForecast).filter(
            PortfolioLRForecast.portfolioid == pf.portfolioid,
            PortfolioLRForecast.yearmonth == pf.yearmonth,
            PortfolioLRForecast.source == sourcename).delete()

        # if the portfolio is in the CI-selected portfolios and has a forecast
        if pf.portfolioid in ciportfolios and pf.forecast != None and pf.forecast != '$0.00':
            # convert forecast to a number
            pfforecast = float(pf.forecast.replace('$','').replace(',',''))
            # for each labor role in the model forecast
            for cipflr in ciportfolios[pf.portfolioid]:
                # create a record with calculated hours in the portfolio labor role forecast
                pflr = PortfolioLRForecast()
                pflr.portfolioid = pf.portfolioid
                pflr.yearmonth = pf.yearmonth
                pflr.laborroleid = cipflr['LaborRoleID']
                pflr.forecasteddollars = pfforecast * cipflr['Pct'] #/ 100
                pflr.forecastedhours = pflr.forecasteddollars / rates[pf.portfolio.clientname][pflr.laborroleid]
                pflr.source = sourcename
                pflr.updateddate = datetime.date.today()
                db.session.add(pflr)
                loglines.append(f"  {pflr.laborroleid} {pflr.forecastedhours} hours")

        db.session.commit()

    return loglines

@app.cli.command('forecast_gsheets')
def forecast_gsheets_cmd():
    forecast_gsheets()

def forecast_gsheets():
    loglines = AdminLog()
    with app.app_context():
        Base.prepare(autoload_with=db.engine, reflect=True)
    loglines.append("Starting Google Sheets Forecast Importer")
    loglines.append("")

    thisyear = datetime.date.today().year
    if datetime.date.today().month > 10:
        thisyear += 1

    source = 'gsheet'

    # gather some lookup data
    loglines.append("gathering lookup data")
    rates = get_clrrates()
    laborroles = db.session.query(LaborRole).all()

    # for each Google Sheet forecast
    for gs in db.session.query(PortfolioLRForecastSheet).all():
        # gs is linked to the CST, but also sometimes to the client name
        # if it's linked to the client name, use that
        clientqueryarg = True
        if gs.clientname != None and gs.clientname != '':
            clientqueryarg = Portfolio.clientname == gs.clientname

        # delete any existing future portfolio labor role forecasts attached to a portfolio that's in the sheet's CST
        db.session.query(PortfolioLRForecast).filter(
            PortfolioLRForecast.portfolioid.in_(db.session.query(Portfolio.id).filter(Portfolio.currcst == gs.cstname)),
            PortfolioLRForecast.yearmonth >= datetime.date.today().replace(day=1),
            clientqueryarg,
            PortfolioLRForecast.source == source).delete(synchronize_session='fetch')

        tabname = gs.tabname
        if tabname == None or tabname == '':
            tabname = 'Export'
        
        # get the sheet
        loglines.append(f"for CST {gs.cstname}, getting sheet {gs.gsheet_url}")
        sheet = getGoogleSheet(gs.gsheet_url)

        # if it has an Export tab, delete it
        worksheets = sheet.worksheets()
        if any(worksheet.title == tabname for worksheet in worksheets):
            loglines.append(f"  deleting {tabname} tab")
            sheet.del_worksheet(sheet.worksheet(tabname))

        worksheets = sheet.worksheets()
        # if it doesn't have an Export tab
        if not any(worksheet.title == tabname for worksheet in worksheets):
            rows = 300000
            # create one, load into a df
            loglines.append(f"  creating {tabname} tab")
            sheet.add_worksheet(tabname, rows=rows, cols=40)
            worksheet = sheet.worksheet(tabname)
            # build the header row
            headerrow = ['Client', 'Portfolio', 'Labor Category', 'Labor Role', 'MSA Rate']
            for month in range(1,13):
                # determine month's name
                monthname = datetime.date(thisyear, month, 1).strftime('%b')
                headerrow.append(f'{monthname} {thisyear} hrs')
                headerrow.append(f'{monthname} {thisyear} $')
            loglines.append("  building dataframe")
            df = pd.DataFrame([["" for _ in range(29)] for _ in range(rows)], columns=headerrow)
            # now, for each clientname and portfolio in the sheet's CST, add a row for each labor role and labor role's rate with the client
            # including summary rows for each client and portfolio, also calculating the total dollars based on the hours and rate
            rownum = 0
            for clientdbrow in db.session.query(Portfolio.clientname).filter(Portfolio.currcst == gs.cstname).distinct().all():
                clientname = clientdbrow[0]
                # client row
                df.iloc[rownum,0] = clientname
                rownum += 1
                loglines.append(f"    row {rownum}")
                for portfoliodbrow in db.session.query(Portfolio.name).filter(Portfolio.currcst == gs.cstname).distinct().all():
                    portfolio = portfoliodbrow[0]
                    #loglines.append(f"  portfolio {clientname} // {portfolio}")
                    # portfolio row
                    df.iloc[rownum,0] = clientname
                    df.iloc[rownum,1] = portfolio
                    portfoliostartrow = rownum
                    rownum += 1
                    #loglines.append(f"    row {rownum}")
                    # labor role rows
                    for lr in laborroles:
                        df.iloc[rownum,0] = clientname
                        df.iloc[rownum,1] = portfolio
                        df.iloc[rownum,2] = lr.categoryname
                        df.iloc[rownum,3] = lr.name
                        if clientname in rates and lr.id in rates.get(clientname):
                            df.iloc[rownum,4] = round(rates[clientname][lr.id],2)
                        for month in range(1,13):
                            hrscolname = openpyxl.utils.get_column_letter((month*2)+4)
                            ratecolname = 'E'
                            df.iloc[rownum,(month*2)+4] = f'=IFERROR({hrscolname}{rownum+2}*{ratecolname}{rownum+2},0)'
                            # this is too slow
                            #setGoogleSheetCellFormat(worksheet, f'{hrscolname}{rownum+2}', { 'backgroundColor': { 'red': 1.0, 'green': 1.0, 'blue': 0.8 } })
                        rownum += 1
                        #loglines.append(f"    row {rownum}")

                    # update the portfolio start row to sum the portfolio's hours and dollars
                    for month in range(1,13):
                        hrscolname = openpyxl.utils.get_column_letter((month*2)+4)
                        dollarscolname = openpyxl.utils.get_column_letter((month*2)+5)
                        df.iloc[portfoliostartrow,(month*2)+3] = f'=SUM({hrscolname}{portfoliostartrow+3}:{hrscolname}{rownum+1})'
                        df.iloc[portfoliostartrow,(month*2)+4] = f'=SUM({dollarscolname}{portfoliostartrow+3}:{dollarscolname}{rownum+1})'

            loglines.append(f"  saving worksheet")
            df = df.fillna('')
            worksheet.update([df.columns.values.tolist()] + df.values.tolist(), raw=False)
            setGoogleSheetCellFormat(worksheet, '1:1', {'textFormat': {'bold': True}})

    # for each forecast in the sheet, delete any labor role forecasts against that client and portfolio that are in the future
    clientstodeletelrforecast = set()
    for row in range(2, worksheet.row_count+1):
        clientname = worksheet.cell(row, 1).value
        if clientname != None and clientname != '':
            clientstodeletelrforecast.add(clientname)

    db.session.query(PortfolioLRForecast).filter(
        PortfolioLRForecast.portfolioid.in_(db.session.query(Portfolio.id).filter(Portfolio.clientname.in_(clientstodeletelrforecast))),
        PortfolioLRForecast.yearmonth >= datetime.date.today().replace(day=1)).delete(synchronize_session='fetch')
    db.session.commit()

    # create a dictionary of portfolio name to portfolio id
    portfolionames = {}
    for row in db.session.query(Portfolio.name, Portfolio.id).all():
        portfolionames[row[0]] = row[1]
    # create a dictionary of labor role name to labor role id
    laborrolenames = {}
    for row in db.session.query(LaborRole.name, LaborRole.id).all():
        laborrolenames[row[0]] = row[1]

    # for each completed forecast in the sheet for this month or a future month, add the forecast to the database (in batches of 100)
    # TODO: finish this
    for row in range(2, worksheet.row_count+1):
        portfolio = worksheet.cell(row, 2).value
        portfolioid = portfolionames.get(portfolio) if portfolio != None and portfolio != '' else None

        laborrole = worksheet.cell(row, 4).value
        laborroleid = laborrolenames.get(laborrole) if laborrole != None and laborrole != '' else None

        if portfolioid != None and laborroleid != None:
            yearmonth = datetime.date(thisyear, thismonth, 1)
            for month in range(1,13):
                hours = worksheet.cell(row, (month*2)+4).value
                if hours != None and hours != '':
                    forecast = PortfolioLRForecast(portfolioid=portfolioid, laborroleid=laborroleid, yearmonth=yearmonth, hours=hours)
                    db.session.add(forecast)
                yearmonth = yearmonth + relativedelta(months=1)
            if row % 100 == 0:
                db.session.commit()
    db.session.commit()

    return loglines

@app.cli.command('replicate_labor_role_hours_day_ratio')
def replicate_labor_role_hours_day_ratio_cmd():
    replicate_labor_role_hours_day_ratio()

def replicate_labor_role_hours_day_ratio():
    loglines = AdminLog()
    loglines.append(f"REPLICATING LABOR ROLE HOURS/DAY RATIO")
    with app.app_context():
        Base.prepare(autoload_with=db.engine, reflect=True)
    # run query against Genome datalake in BQ
    loglines.append(f"  running query")
    client = bigquery.Client()
    query_job = client.query("""
select LaborRoleID, LaborRole, avg(Hours) / round(avg(HeadCount),2) as HoursPerDay, round(avg(HeadCount),2) as HeadCount
from (
  select uwd.YearMonthDay, uwd.LaborRoleID, uwd.LaborRole, sum(Hours) as Hours, sum(HeadCount) as HeadCount
  from 
  (
    select uwd.YearMonthDay, uwd.LaborRoleID, uwd.LaborRole, count(distinct UserID) as HeadCount
    from GenomeDW.DUserWorkDay uwd 
    where uwd.EmployeeTypeID = 'PM' --and uwd.YearMonthDay = '2022-08-19' and uwd.LaborRoleID = 'CLSRVDIR'
		and uwd.BillablePercent >= 0.5
		and uwd.IsAtWork = true and uwd.IsActive = true
    group by uwd.YearMonthDay, uwd.LaborRoleID, uwd.LaborRole
  ) uwd 
  inner join 
  (
    select cast(fa.ActualDateTime as date) as YearMonthDay, sum(fa.Hours) as Hours, fa.LaborRole
    from `genome-datalake-prod.GenomeDW.F_Actuals` fa
    inner join `genome-datalake-prod.GenomeDW.DUser` u on fa.Employee = u.UserID
    inner join `genome-datalake-prod.GenomeDW.Project` p on fa.Project = p.Project
    where u.EmployeeTypeID = 'PM'
    and fa.ActualDateTime >= cast(date_sub(current_date('EST5EDT'), INTERVAL 4 MONTH) AS TIMESTAMP)
		and p.billable=true
    group by fa.LaborRole, cast(fa.ActualDateTime as date)
  ) a on uwd.YearMonthDay = cast(a.YearMonthDay as date) and uwd.LaborRoleID = a.LaborRole
  where extract(dayofweek from uwd.YearMonthDay) between 2 and 6
  group by uwd.YearMonthDay, uwd.LaborRoleID, uwd.LaborRole
) a
group by LaborRoleID, LaborRole
order by HoursPerDay desc
    """)
    results = query_job.result()

    # for each row in the results, update the database
    for row in results:
        ratio = db.session.query(LaborRoleHoursDayRatio).filter(LaborRoleHoursDayRatio.laborroleid == row.LaborRoleID).first()
        if ratio == None:
            loglines.append(f"  adding {row.LaborRole}")
            ratio = LaborRoleHoursDayRatio()
            db.session.add(ratio)
        else:
            loglines.append(f"  updating {row.LaborRole}")

        ratio.laborroleid = row.LaborRoleID
        ratio.name = row.LaborRole
        ratio.hoursperday = round(row.HoursPerDay,2)
        ratio.headcount = round(row.HeadCount,2)

    loglines.append("  committing")
    db.session.commit()
    return loglines





admin.add_view(ForecastAdminView(name='Forecast Processing', category='Forecasts'))



